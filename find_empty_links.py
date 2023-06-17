import openpyxl as xl

wb = xl.load_workbook('links.xlsx')
sheet = wb['links']

# for each row
for row in range(2, sheet.max_row + 1):
    # if an empty row - stop
    if sheet.cell(row, 1).value is None:
        break
    try:
        llink = sheet.cell(row, 1).hyperlink.target
    except Exception as e:
        sheet.cell(row, 2).value = f'{str(e)}'
    if llink in (None, ''):
        sheet.cell(row, 2).value = f'link is empty'

   
try:
    wb.save('links.xlsx')
except Exception as e:
    print(f'Cannot save the excel file: {str(e)}')
