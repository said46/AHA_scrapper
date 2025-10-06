import openpyxl as xl
import os
from my_utils import message_box 

os.system('cls')


def save_excel_file(file_name) -> None:
    try:
        wb.save(file_name)
    except Exception as e:
        print(f'Cannot save the excel file: {str(e)}')


xlsx_name = '_links.xlsx'
pdfs_folder_path = os.getcwd() + '\\_get_pdfs'

wb = xl.load_workbook(xlsx_name)
sheet = wb['links']

list_of_file_names = [_ for _ in os.listdir(pdfs_folder_path)]
list_of_file_names.sort(key=lambda xa: os.path.getctime(os.path.join(pdfs_folder_path, xa)))

doc_list = list() # contains (row, doc_number) for each doc
for row in range(2, sheet.max_row + 1):
    doc_number = sheet.cell(row, 1).value        
    if doc_number is None:
        break            
    if sheet.cell(row, 2).value == "downloaded":
        doc_list.append((row, doc_number))

if len(doc_list) != len(list_of_file_names):
    message_box('length of doc_list is not equal to length of list_of_file_names', 'Ok!', 0)    
    wb.close()
    quit()

for dn, fn in zip(doc_list, list_of_file_names):
    row = dn[0]
    doc_number = dn[1]    
    new_name = os.path.join(pdfs_folder_path, doc_number + '.pdf')
    try:
        ####################################################################
        # IF NEW FILENAME CONTAINS # LINKS IN EXCEL WILL NOT WORK!!!!!!
        ####################################################################
        last_slash_pos = new_name.rfind("\\")
        print(f"row {row}: Renaming {fn} to {new_name[last_slash_pos + 1:]} ...")
        os.rename(os.path.join(pdfs_folder_path, fn), os.path.join(pdfs_folder_path, new_name))
        sheet.cell(row, 2).value = "downloaded and renamed"
    except Exception as e:
        sheet.cell(row, 2).value = f"{str(e)}"
        continue
    finally:
        save_excel_file(xlsx_name)

wb.close()

message_box('Ok?', 'Ok!', 0)