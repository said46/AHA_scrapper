from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from my_utils import message_box 
import openpyxl as xl
import os


def save_excel_file(file_name) -> None:
    try:
        wb.save(file_name)
    except Exception as e:
        print(f'Cannot save the excel file: {str(e)}')


xlsx_name = '_links.xlsx'
pdfs_folder_path = os.getcwd() + '\\_get_pdfs'
options = EdgeOptions()
options.add_experimental_option('prefs', {
                                            "download.prompt_for_download": False,
                                            "plugins.always_open_pdf_externally": True,
                                            "download.default_directory": pdfs_folder_path, 
                                            "download.directory_upgrade": True
                                         })

edgeBrowser = webdriver.Edge(r"msedgedriver.exe", options=options)
edgeBrowser.maximize_window()

wb = xl.load_workbook(xlsx_name)
sheet = wb['links']

# for each row
count = 0 # чтобы качать пачками (когда около 2000 файлов, что-нибудь да не так идёт)
for row in range(2, sheet.max_row + 1):
    # if an empty row - stop
    if sheet.cell(row, 1).value is None:
        break        

    if sheet.cell(row, 2).value is not None:
        continue
    
    # если размер пачки равен ..., то заканчиваем (чтобы переименовать другим скриптом и перенести переименованные файлы в другую папку)
    if count == 300:
        message_box(f'limit={count}', 'Ok', 0)
        edgeBrowser.quit()
        save_excel_file(xlsx_name)
        wb.close()        
        quit()

    try:
        llink = sheet.cell(row, 1).hyperlink.target
    except Exception as e:
        sheet.cell(row, 2).value = f"{str(e)}" 
        save_excel_file(xlsx_name)
        continue
    
    doc_number = sheet.cell(row, 1).value
    try:
        edgeBrowser.get(llink)
    except Exception as e:
        sheet.cell(row, 2).value = f"{str(e)}"        
        save_excel_file(xlsx_name)
        continue

    sheet.cell(row, 2).value = f'downloaded'
    count += 1
    save_excel_file(xlsx_name)

message_box('Ok?', 'Ok!', 0)
edgeBrowser.quit()
wb.close()
