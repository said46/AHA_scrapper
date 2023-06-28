from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import InvalidSelectorException
from selenium.webdriver.edge.options import Options as EdgeOptions
import ctypes
import os
import openpyxl as xl


def message_box(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


#  Styles:
#  0 : OK *** #  1 : OK | Cancel *** 2 : Abort | Retry | Ignore *** 3 : Yes | No | Cancel ***  
#  4 : Yes | No *** 5 : Retry | Cancel *** 6 : Cancel | Try Again | Continue


def prepare_tag_for_search(tag: str, asset: str) -> str:
    # temp begin
    # if tag[:11] == "%%O052LI001":
    #     result = '*' + asset + '-' + tag[3:6] + '-LI-' + tag[8:12]
    #     return result
    # temp end
    if tag[-3] == '_':
        tag = tag[:-3]
    if tag[0:2] == '%%':
        tag = tag[3:]
    if tag[-1].isalpha():
        last_part_beginning_index = -4
    else:
        last_part_beginning_index = -3
    middle_part = tag[3:last_part_beginning_index]
    if middle_part[1:] in ('ICA', 'IA', 'I', 'IC') and middle_part != 'HIC':
        middle_part = middle_part[0] + 'T'
    if middle_part in ('XZGV', 'XZGOC'):
        middle_part = 'XZV'
    if middle_part in ('XGV', 'XGOC'):
        middle_part = 'XV'
    if middle_part in ('HGV', 'HIC', 'HIQ'):
        middle_part = 'HV'
    if middle_part in ('PDI', 'PDIA', 'PDIC', 'PDICA'):
        middle_part = 'PDT'
    if middle_part == 'PDSH':
        middle_part = 'PD*'
    # temp begin
    # if middle_part == 'XX':
    #     middle_part = 'XA'          
    # if middle_part == 'XS':
    #     middle_part = 'XA'         
    # temp end
    # temp begin 
    # if middle_part == 'XZGC':
    #    middle_part = 'XGC'        
    # if middle_part == 'XZGO':
    #    middle_part = 'XGO'          
    # temp end
    result = '*' + asset + '-' + tag[:3] + '-' + middle_part + '-' + tag[last_part_beginning_index:]
    return result



excel_columns = {"tag": 1, "desc": 2, "card": 3, "folder_link": 4, "doc_link": 5, "tag_prep": 6, "result": 12}
search_excel_name = '_search.xlsx'

os.system("cls")

options = EdgeOptions()
options.add_argument("start-maximized")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
edgeBrowser = webdriver.Edge(r"msedgedriver.exe", options=options)

try:
    edgeBrowser.get('http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/default.asp?AHAContextID=1')
except Exception as e:
    print(f"{str(e)}, aborting the program")
    edgeBrowser.quit()
    quit()

key0 = "4000"

try:
    wb = xl.load_workbook(search_excel_name)
except Exception as e:
    print(f'Cannot open the excel file: {str(e)}')
    quit()

sheet = wb['Input Data']
for count, row in enumerate(range(2, sheet.max_row + 1)):
    # if an empty row - stop
    if sheet.cell(row, excel_columns["tag"]).value in (None, ''):
        break

    # stop for test purposes
    if count == 50000:
        break

    # no need to process the already processed rows
    if sheet.cell(row, excel_columns["folder_link"]).value not in (None, '') or \
            sheet.cell(row, excel_columns["result"]).value not in (None, ''):
        continue

    sheet.cell(row, excel_columns["doc_link"]).value = ''
    sheet.cell(row, excel_columns["folder_link"]).value = ''
    sheet.cell(row, excel_columns["tag_prep"]).value = ''
    sheet.cell(row, excel_columns["result"]).value = ''

    card = sheet.cell(row, excel_columns["card"]).value

    key0_param = key0[:-1]

    tag_prepared_for_search = prepare_tag_for_search(sheet.cell(row, excel_columns["tag"]).value, key0_param)
    sheet.cell(row, excel_columns["tag_prep"]).value = tag_prepared_for_search
    sheet.cell(row, excel_columns[
        "tag_prep"]).hyperlink = "http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/search/Search.asp?" \
                                 "obj_type_id=4" \
                                 "&obj_type_name=Tag&cls_obj_name=&" \
                                 f"obj_name={tag_prepared_for_search}&obj_desc=*&key0={key0}"
    sheet.cell(row, excel_columns["tag_prep"]).style = "Hyperlink"

    # 2 tries to search, one with desc == '*', if no success, with desc == first letter of tag desc + '*'
    obj_descs = list()
    obj_descs.append('*')
    if sheet.cell(row, excel_columns["desc"]).value not in ('', None):
        obj_descs.append(sheet.cell(row, excel_columns["desc"]).value[0] + '*')

    node_id = ''
    for obj_desc in obj_descs:
        llink = 'http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/treeview/tree.asp?Option=ObjectSearch' \
                '&obj_type_id=4' \
                '&obj_type_name=Tag' \
                f'&cls_obj_name=&obj_name={tag_prepared_for_search}' \
                f'&obj_desc={obj_desc}' \
                f'&key0={key0}'
        if obj_desc == '*':
            sheet.cell(row, excel_columns["tag"]).hyperlink = llink
        sheet.cell(row, excel_columns["tag"]).style = "Hyperlink"
        edgeBrowser.get(llink)

        edgeBrowser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        node_id = ''
        element_xpath = "//a[text()='Document(s)']"
        try:
            node_id = edgeBrowser.find_element(By.XPATH, element_xpath).get_attribute(name='id')
        except NoSuchElementException:
            sheet.cell(row, excel_columns["result"]).value = "Node 'Document(s)' is not found (number of search results is 0 or >1)"

        if node_id != '':
            llink = "http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/treeview/tree.asp?Option=ClickNode" \
                    "&ScrollPosX=0" \
                    "&ScrollPosY=0" \
                    f"&TreeNodeID={node_id}" \
                    "&IsObjectNode=False"
            edgeBrowser.get(llink)
            break
        else:
            continue

    # Node 'Document(s)' is not found
    if node_id == '':
        continue

    # 2 tries to find doc_number, one with 'LOOP', another with 'SEGMENT' if ALF111
    xpaths = []
    xpaths.append("//a[contains(translate(., 'instrument loop', 'INSTRUMENT LOOP'), 'INSTRUMENT LOOP') and "    
                  "not(contains(translate(., 'typical', 'TYPICAL'), 'TYPICAL')) and "                  
                  "not(contains(translate(., 'fire', 'FIRE'), 'FIRE')) and "
                  "not(contains(translate(., 'index', 'INDEX'), 'INDEX')) and " 
                  "not(contains(translate(., 'punch', 'PUNCH'), 'PUNCH'))]")
    if card == "ALF111":
        xpaths.append("//a[contains(translate(., 'segment', 'SEGMENT'), 'SEGMENT')]")                  

    doc_number = ''

    for element_xpath in xpaths:
        try:
            elem = edgeBrowser.find_element(By.XPATH, element_xpath)
            doc_number = elem.text.split(',')[0]
        except NoSuchElementException:
            sheet.cell(row, excel_columns["result"]).value = "Node with text containing 'LOOP' or 'SEGMENT" \
                                                             "(case insensitive) is not found"
            continue
        except InvalidSelectorException:
            sheet.cell(row, excel_columns["result"]).value = f'Invalid XPATH={element_xpath} expression'
            continue

    if doc_number == '':
        continue

    llink = "http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/treeview/tree.asp?Option=ObjectSearch&" \
            "obj_type_id=7&obj_type_name=Document&cls_obj_name=AhaQryStdRel.FindObjectWoRev%28%27Document%27%2C%27document+issue+date%27%29&" \
            f"obj_name={doc_number}*&obj_desc=*"
    llink = llink.replace('#', '%23')
    edgeBrowser.get(llink)

    element_xpath = "//a[contains(translate(., 'loop', 'LOOP'), 'LOOP') or contains(translate(., 'segment', 'SEGMENT'), 'SEGMENT')]"
    try:
        node_id = edgeBrowser.find_element(By.XPATH, element_xpath).get_attribute(name='id')
    except NoSuchElementException:
        sheet.cell(row, excel_columns["result"]).value = \
            "While searching by doc_number, node with text containing 'LOOP' (case insensitive) is not found"
        continue
    except InvalidSelectorException:
        sheet.cell(row, excel_columns["result"]).value = f'Invalid XPATH={element_xpath} expression'
        continue

    llink = f"http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/relationsandmethods.asp?TreeNodeID={node_id}&ScrollPosX=0&ScrollPosY=0"
    edgeBrowser.get(llink)

    element_xpath = "//a[text()='Jump in Unica compound document']"
    try:
        llink = edgeBrowser.find_element(By.XPATH, element_xpath).get_attribute("href")
    except NoSuchElementException:
        sheet.cell(row, excel_columns[
            "result"]).value = "Node with text equal to 'Jump in Unica compound document' is not found"
        continue
    except InvalidSelectorException:
        sheet.cell(row, excel_columns["result"]).value = f'Invalid XPATH={element_xpath} expression'
        continue

    edgeBrowser.get(llink)

    window_after = edgeBrowser.window_handles[1]
    edgeBrowser.switch_to.window(window_after)

    folder_link = edgeBrowser.current_url
    sheet.cell(row, excel_columns["folder_link"]).hyperlink = folder_link
    sheet.cell(row, excel_columns["folder_link"]).value = 'UNICA folder link'
    sheet.cell(row, excel_columns["folder_link"]).style = "Hyperlink"

    element_xpath = "//a[contains(translate(., 'loop diagram', 'LOOP DIAGRAM'), 'LOOP DIAGRAM') or " \
                    "contains(translate(., 'segment', 'SEGMENT'), 'SEGMENT')]"
    try:
        doc_link = edgeBrowser.find_element(By.XPATH, element_xpath).get_attribute("href")
        sheet.cell(row, excel_columns["doc_link"]).hyperlink = doc_link
        sheet.cell(row, excel_columns["doc_link"]).value = doc_number
        sheet.cell(row, excel_columns["doc_link"]).style = "Hyperlink"
    except NoSuchElementException:
        sheet.cell(row, excel_columns[
            "result"]).value = "Node with text containing 'LOOP DIAGRAM' (case insensitive) is not found"
        continue
    except InvalidSelectorException:
        sheet.cell(row, excel_columns["result"]).value = f'Invalid XPATH={element_xpath} expression'
        continue

    # save Excel file every 10 rows
    if count % 10 == 0 and count > 0:
        try:
            wb.save(search_excel_name)
        except Exception as e:
            print(f'Cannot save the excel file: {str(e)}')

try:
    wb.save(search_excel_name)
except Exception as e:
    print(f'Cannot save the excel file: {str(e)}')

wb.close()
message_box('Ok?', 'Ok!', 0)
edgeBrowser.quit()
